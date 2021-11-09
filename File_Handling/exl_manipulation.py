# -------------------------------------------------------------------------------
# Name:        exl_manipulation.py
# Project:     Python_Programming_Labs
#
# Author:      Vitaliy Baseckas
#
# Created:     09.11.2021
# Copyright:   (c) Vitaliy Baseckas 2021
# Licence:     <your licence>
# -------------------------------------------------------------------------------
"""
Description
----------
Parsing Excel files with openpyxl
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import colors, Font, Color
from openpyxl.comments import Comment
from openpyxl import load_workbook
list_of_rows=[]

#workbook = load_workbook('/tmp/data.xlsx')
#first_sheet = workbook.get_sheet_names()[0]
#worksheet = workbook.get_sheet_by_name(first_sheet)

#for row in worksheet.iter_rows():
#    for cell in row:
#        if cell.comment:
#            print(cell.comment.text)

def parse_xl_file(file_name):
    """
    Parse an excel file using openpyxl
    :param file_name:
    Name of the file to be parsed
    :return:
    None
    """
    print("Hi")
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb['Sheet1']
    #print(ws['A1'].value)
    #print(ws['A2'].value)
    # Iterate through the excel file by row
    for i in range(1, ws.max_row + 1):
        #print("\n")
        #print("Row ", i, " data :")
        font_1 = Font(color=colors.BLUE, italic=False, bold=True)
        #cell_obj = ws.cell(row=i, column=j)
        #print(cell_obj.value, end=" ")
        #cell_obj.font = font_1
    # iterate thorugh the columns
        list_of_cells=[]
        for j in range(1, ws.max_column + 1):
            cell_obj = ws.cell(row=i, column=j)
            list_of_cells.append(cell_obj.value)
            cell_obj.font = font_1
            #print(cell_obj.value, end=" ")
        list_of_rows.append(list_of_cells)

def new_xls_file():
    font_1 = Font(color=colors.BLUE, italic=False, bold=True)
    wb1=Workbook()
    ws1 = wb1.active
    ws1 = wb1.create_sheet("Mysheet", 0)  # insert at first position
    print(list_of_rows)
    l=0
    for data in list_of_rows:
        i=0
        for c in data:
            my_cell=ws1.cell(row=l+1, column=i+1)
            my_cell.value = c
            my_cell.font=font_1
            comments = Comment(text='Added comments', author='Vitaliy Baseckas') #formating comments
            ws1.cell(row=l+1, column=1).comment=comments #asigning comments
            if my_cell.comment: #checking if the cell note:
                print(my_cell.comment.text) #print comments if there are such
                #pass
            else:
                pass
            i+=1
        l+=1
    wb1.save('test.xlsx')

if __name__ == "__main__":
    parse_xl_file('Financial Sample.xlsx')
    new_xls_file()
